from io import BytesIO
from typing import Any, Dict, Iterable, List, Optional, Tuple

from openpyxl import load_workbook

NBSP = "\u00a0"


def _norm_text(s: str) -> str:
    s = s.replace(NBSP, " ").replace("\t", " ").strip()
    s = " ".join(s.split())
    return s


def _norm_header(h: Any) -> str:
    if h is None:
        return ""
    if isinstance(h, str):
        return _norm_text(h).casefold()
    return str(h).strip().casefold()


def _clean(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, str):
        s = _norm_text(v)
        if not s or s.casefold() in {"nan", "none", "null"}:
            return None
        return s
    return v


def norm_wav_key(v: Any) -> Optional[str]:
    v = _clean(v)
    if v is None:
        return None

    normalized = _norm_text(str(v).replace("\\", "/"))
    if "/" in normalized:
        normalized = normalized.rsplit("/", 1)[-1].strip()
    return normalized or None


def _iter_table_rows(ws, table) -> Iterable[Tuple[Any, ...]]:
    for row in ws[table.ref]:
        yield tuple(cell.value for cell in row)


HEADER_ALIASES = {
    "wav_name": {"wav_name", "wav name", "wav", "wavname", "filename", "file_name"},
    "csv_name": {"csv_name", "csv name", "csv", "csvname"},
    "conversation_id": {"conversation_id", "conversation id", "conversationid"},
    "start_time": {"start_time", "start time"},
    "end_time": {"end_time", "end time"},
    "duration": {"duration", "duracion", "duración"},
    "xlsx_name": {"xlsx_name", "xlsx name"},
    "consecutivo": {"consecutivo", "consecutive"},
    "fecha_ocurrencia": {"fecha_ocurrencia", "fecha ocurrencia"},
    "tiempo_ocurrencia": {"tiempo_ocurrencia", "tiempo ocurrencia"},
    "activo_herope": {"activo_herope", "activo herope"},
    "tipo_reporte": {"tipo_reporte", "tipo reporte"},
    "tipo_movimiento": {"tipo_movimiento", "tipo movimiento"},
    "denominación_causa e-bitacora": {
        "denominación_causa e-bitacora",
        "denominacion_causa e-bitacora",
        "denominación causa e-bitacora",
        "denominacion causa e-bitacora",
    },
}

PUBLIC_FIELD_MAP = {
    "Wav_Name": "wav_name",
    "Csv_Name": "csv_name",
    "Conversation_ID": "conversation_id",
    "Start_Time": "start_time",
    "End_Time": "end_time",
    "Duration": "duration",
    "Xlsx_Name": "xlsx_name",
    "Consecutivo": "consecutivo",
    "Fecha_Ocurrencia": "fecha_ocurrencia",
    "Tiempo_Ocurrencia": "tiempo_ocurrencia",
    "Activo_Herope": "activo_herope",
    "Tipo_reporte": "tipo_reporte",
    "Tipo_Movimiento": "tipo_movimiento",
    "Denominación_causa E-Bitacora": "denominación_causa e-bitacora",
}

_CANONICAL_HEADER_BY_ALIAS = {
    alias: canonical
    for canonical, aliases in HEADER_ALIASES.items()
    for alias in aliases
}


def _canonical_header(h_norm: str) -> str:
    return _CANONICAL_HEADER_BY_ALIAS.get(h_norm, h_norm)


def _read_table_blocks(ws) -> List[List[Tuple[Any, ...]]]:
    tables = list(ws.tables.values()) if getattr(ws, "tables", None) else []
    if not tables:
        return []

    selected = _select_best_table_block(ws, tables)
    return [selected] if selected else []


def _read_full_sheet_block(ws) -> List[List[Tuple[Any, ...]]]:
    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        return []
    return [all_rows]


def _get_candidate_blocks(ws) -> List[List[Tuple[Any, ...]]]:
    table_blocks = _read_table_blocks(ws)
    if table_blocks:
        return table_blocks
    return _read_full_sheet_block(ws)


def _select_best_table_block(ws, tables) -> Optional[List[Tuple[Any, ...]]]:
    fallback = None
    for table in tables:
        block = list(_iter_table_rows(ws, table))
        if not block:
            continue

        normalized_headers = [_norm_header(h) for h in block[0]]
        if "wav_name" in normalized_headers:
            return block

        if fallback is None:
            fallback = block
    return fallback


def _is_usable_block(rows: List[Tuple[Any, ...]]) -> bool:
    return bool(rows)


def _split_header_and_data(rows: List[Tuple[Any, ...]]) -> Tuple[Tuple[Any, ...], List[Tuple[Any, ...]]]:
    return rows[0], rows[1:]


def _build_headers_canonical(raw_headers: Tuple[Any, ...]) -> List[str]:
    return [_canonical_header(_norm_header(header)) for header in raw_headers]


def _is_empty_row(values: Tuple[Any, ...]) -> bool:
    return not values


def _build_row(headers_canon: List[str], values: Tuple[Any, ...]) -> Dict[str, Any]:
    row: Dict[str, Any] = {}
    for index, canonical_header in enumerate(headers_canon):
        if not canonical_header:
            continue
        value = _clean(values[index] if index < len(values) else None)
        if value is not None:
            row[canonical_header] = value
    return row


def _merge_manifest_row(manifest: Dict[str, Dict[str, Any]], row: Dict[str, Any]) -> None:
    key = norm_wav_key(row.get("wav_name"))
    if not key:
        return

    current = manifest.get(key, {})
    for field, value in row.items():
        if value is not None:
            current[field] = value
    manifest[key] = current


def _to_public_manifest(manifest: Dict[str, Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    final: Dict[str, Dict[str, Any]] = {}
    for key, values in manifest.items():
        mapped = {
            public_key: values.get(source_key)
            for public_key, source_key in PUBLIC_FIELD_MAP.items()
        }
        final[key] = {field: value for field, value in mapped.items() if value is not None}
    return final


def _transform_block_into_manifest_rows(
    rows: List[Tuple[Any, ...]],
    manifest: Dict[str, Dict[str, Any]],
) -> None:
    if not _is_usable_block(rows):
        return

    raw_headers, data_rows = _split_header_and_data(rows)
    headers_canon = _build_headers_canonical(raw_headers)

    for values in data_rows:
        if _is_empty_row(values):
            continue
        row = _build_row(headers_canon, values)
        _merge_manifest_row(manifest, row)


def parse_xlsx_manifest(xlsx: BytesIO) -> Dict[str, Dict[str, Any]]:
    workbook = load_workbook(xlsx, data_only=True)
    worksheet = workbook.active

    candidate_blocks = _get_candidate_blocks(worksheet)
    if not candidate_blocks:
        return {}

    manifest: Dict[str, Dict[str, Any]] = {}
    for rows in candidate_blocks:
        _transform_block_into_manifest_rows(rows, manifest)

    return _to_public_manifest(manifest)
